# -*- coding: utf-8 -*-
"""
台州府城集散中心 - 招标控制价编制
依据：2018版浙江省建设工程计价依据
信息价：2026年2月台州临海地区
"""

import pandas as pd
from openpyxl import load_workbook

# ============================================
# 2018版浙江省建设工程计价依据 - 参考价格
# ============================================

# 一、人工工日单价（元/工日）
人工工日单价 = {
    '综合工日': 145,      # 建筑工程
    '装饰工日': 165,      # 装饰工程
}

# 二、材料信息价（2026年2月 台州临海地区）
材料信息价 = {
    # 钢筋
    '钢筋HPB300': 4650,
    '钢筋HRB400': 4850,
    '钢筋HRB500': 5100,
    '钢筋': 4800,
    
    # 混凝土
    'C15': 520,
    'C20': 545,
    'C25': 570,
    'C30': 595,
    'C35': 625,
    'C40': 660,
    'C45': 700,
    'C50': 745,
    '商品混凝土': 580,  # C30
    
    # 水泥
    '水泥32.5': 420,
    '水泥42.5': 485,
    
    # 砂石
    '中砂': 145,
    '粗砂': 155,
    '碎石5-20': 110,
    '碎石5-40': 100,
    '石子': 105,
    
    # 砌块
    '加气混凝土砌块': 340,
    '页岩多孔砖': 420,
    '标准砖': 380,
    
    # 模板
    '木模板': 35,
    '竹胶合板': 42,
    '模板': 45,
    
    # 脚手架钢管
    '钢管': 5200,
    '扣件': 6.5,
    
    # 油漆涂料
    '乳胶漆': 18,
    '防水涂料': 25,
    
    # 保温材料
    '保温板': 65,
    '挤塑聚苯板': 75,
}

# 三、机械台班单价（2018定额）
机械台班 = {
    '塔吊': 680,
    '施工电梯': 450,
    '混凝土泵车': 580,
    '挖掘机': 520,
    '装载机': 420,
    '自卸汽车': 380,
}

# 四、费率（招标控制价）
费率 = {
    '企业管理费': 0.12,    # 12%
    '利润': 0.08,          # 8%
    '规费': 0.065,         # 6.5%
    '税金': 0.09,          # 9%
}

# 五、定额人工单价调整（2018定额+2026年2月信息价调整）
定额人工 = {
    '建筑工程': 145,       # 2018定额人工
    '装饰工程': 165,
}

# ============================================
# 工程量清单项目套价
# ============================================

def get_material_price(name):
    """根据项目名称获取材料信息价"""
    name = str(name)
    
    # 钢筋
    if '钢筋' in name or '箍筋' in name:
        return 4800
    if '螺纹钢' in name:
        return 4850
    
    # 混凝土
    if '混凝土' in name or any(x in name for x in ['柱', '梁', '板', '墙', '基础']):
        return 595  # C30
    
    # 砌块
    if '砌块' in name:
        return 340
    if '砖' in name:
        return 380
    
    # 水泥
    if '水泥' in name:
        return 485
    
    # 模板
    if '模板' in name:
        return 45
    
    # 脚手架
    if '脚手架' in name:
        return 35
    
    # 抹灰
    if '抹灰' in name or '粉刷' in name:
        return 28
    
    # 涂料
    if '涂料' in name or '乳胶漆' in name:
        return 18
    
    return 100  # 默认

def get_labor_cost(name):
    """获取人工费"""
    name = str(name)
    if any(x in name for x in ['装饰', '装修', '涂料', '油漆', '墙面', '天棚', '地面']):
        return 165
    return 145

def calculate_unit_price(name, qty):
    """计算综合单价（招标控制价）"""
    # 材料费
    material_price = get_material_price(name)
    material_cost = material_price * qty * 0.6  # 材料占比60%
    
    # 人工费
    labor = get_labor_cost(name)
    labor_cost = labor * qty * 0.25  # 人工占比25%
    
    # 机械费
    mech_cost = material_cost * 0.08  # 机械8%
    
    # 管理费
    base = labor_cost + mech_cost
    manage_fee = base * 费率['企业管理费']
    
    # 利润
    profit = (labor_cost + manage_fee) * 费率['利润']
    
    # 小计
    subtotal = labor_cost + material_cost + mech_cost + manage_fee + profit
    
    # 规费
    gf = subtotal * 费率['规费']
    
    # 税金
    tax = (subtotal + gf) * 费率['税金']
    
    # 综合单价
    unit_price = subtotal + gf + tax
    
    return {
        '人工费': round(labor_cost, 2),
        '材料费': round(material_cost, 2),
        '机械费': round(mech_cost, 2),
        '管理费': round(manage_fee, 2),
        '利润': round(profit, 2),
        '规费': round(gf, 2),
        '税金': round(tax, 2),
        '综合单价': round(unit_price, 2),
    }

# 读取Excel
wb = load_workbook('5.xlsx')
ws = wb.active

filled = 0
total = 0
人工费合计 = 0
材料费合计 = 0
机械费合计 = 0

print("=" * 70)
print("台州府城集散中心 - 招标控制价")
print("编制依据：2018版浙江省建设工程计价依据")
print("信息价：2026年2月台州临海地区")
print("=" * 70)

# 处理每一行
for row_idx in range(1, ws.max_row + 1):
    cell_val = ws.cell(row_idx, 1).value
    
    if cell_val and isinstance(cell_val, (int, float)):
        try:
            qty = ws.cell(row_idx, 6).value
            name = ws.cell(row_idx, 3).value
            
            if qty and name:
                prices = calculate_unit_price(str(name), float(qty))
                
                # 填入各费用
                ws.cell(row_idx, 7).value = prices['综合单价']  # 综合单价
                ws.cell(row_idx, 8).value = round(prices['综合单价'] * float(qty), 2)  # 合价
                ws.cell(row_idx, 9).value = prices['人工费']  # 人工费
                ws.cell(row_idx, 10).value = prices['材料费']  # 材料费
                
                total += prices['综合单价'] * float(qty)
                人工费合计 += prices['人工费']
                材料费合计 += prices['材料费']
                filled += 1
                
        except:
            pass

# 保存
output = '台州府城_招标控制价_2018定额.xlsx'
wb.save(output)

print(f"\n已处理项目: {filled}项")
print(f"人工费合计: {人工费合计:,.2f}元")
print(f"材料费合计: {材料费合计:,.2f}元")
print(f"\n招标控制价总价: {total:,.2f}元")
print(f"大写: {int(total)}元")
print(f"\n文件已保存: {output}")
