# -*- coding: utf-8 -*-
import pandas as pd
from openpyxl import load_workbook

wb = load_workbook('5.xlsx')
ws = wb.active

print("=== 检查前20项的工程量 ===")
for row_idx in range(1, 30):
    cell_val = ws.cell(row_idx, 1).value
    if cell_val and isinstance(cell_val, (int, float)):
        try:
            name = ws.cell(row_idx, 3).value
            unit = ws.cell(row_idx, 5).value
            qty = ws.cell(row_idx, 6).value
            if qty and name:
                print(f"{cell_val}. {str(name)[:20]:20s} 单位:{unit} 工程量:{qty}")
        except:
            pass
