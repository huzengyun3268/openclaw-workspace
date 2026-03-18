# -*- coding: utf-8 -*-
import pandas as pd
from openpyxl import load_workbook

# 进一步调低的单价
定额单价 = {
    '混凝土柱': 55, '混凝土梁': 53, '混凝土板': 52, '混凝土墙': 58,
    '混凝土基础': 50, '混凝土楼梯': 55,
    '钢筋': 450,
    '砌块': 28, '砖': 25,
    '模板': 4.5, '脚手架': 2.2,
    '抹灰': 1.8, '墙面': 2.2, '地面': 2.0, '天棚': 1.5,
    '涂料': 1.2, '防水': 2.3,
    'default': 18,
}

def get_price(name):
    name = str(name)
    if '钢筋' in name: return 定额单价['钢筋']
    if '柱' in name: return 定额单价['混凝土柱']
    if '梁' in name: return 定额单价['混凝土梁']
    if '板' in name: return 定额单价['混凝土板']
    if '墙' in name: return 定额单价['混凝土墙']
    if '基础' in name: return 定额单价['混凝土基础']
    if '楼梯' in name: return 定额单价['混凝土楼梯']
    if '砌' in name: return 定额单价['砌块']
    if '模板' in name: return 定额单价['模板']
    if '脚手架' in name: return 定额单价['脚手架']
    if '抹灰' in name: return 定额单价['抹灰']
    if '墙面' in name: return 定额单价['墙面']
    if '地面' in name: return 定额单价['地面']
    if '天棚' in name: return 定额单价['天棚']
    if '涂料' in name: return 定额单价['涂料']
    if '防水' in name: return 定额单价['防水']
    return 定额单价['default']

wb = load_workbook('5.xlsx')
ws = wb.active

清单 = 0
for row in range(1, ws.max_row + 1):
    val = ws.cell(row, 1).value
    if val and isinstance(val, (int, float)):
        try:
            qty = ws.cell(row, 6).value
            name = ws.cell(row, 3).value
            if qty and name and float(qty) > 0:
                p = get_price(str(name))
                t = float(qty) * p
                ws.cell(row, 7).value = p
                ws.cell(row, 8).value = round(t, 2)
                ws.cell(row, 9).value = round(t*0.2, 2)
                ws.cell(row, 10).value = round(t*0.65, 2)
                清单 += t
        except: pass

企业管理费 = 清单 * 0.12
利润 = 清单 * 0.08
规费 = (清单+企业管理费+利润) * 0.065
税金 = (清单+企业管理费+利润+规费) * 0.09
总价 = 清单 + 企业管理费 + 利润 + 规费 + 税金

wb.save('台州府城_招标控制价.xlsx')

print(f"清单合价: {清单:,.0f}元")
print(f"企业管理费: {企业管理费:,.0f}元")
print(f"利润: {利润:,.0f}元")
print(f"规费: {规费:,.0f}元")
print(f"税金: {税金:,.0f}元")
print(f"\n招标控制价: {总价:,.0f}元")
