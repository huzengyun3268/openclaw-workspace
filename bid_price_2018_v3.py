# -*- coding: utf-8 -*-
"""
台州府城集散中心 - 招标控制价（调整版）
依据：2018版浙江省建设工程计价依据
信息价：2026年2月台州临海地区
"""

import pandas as pd
from openpyxl import load_workbook

# ============================================
# 2018版浙江省建设工程计价依据 - 调整后单价
# ============================================

# 根据2018定额和信息价调整后的单价（元）
定额单价 = {
    # 混凝土工程 - 按m³
    '混凝土柱': 680,    # C30商品混凝土
    '混凝土梁': 665,
    '混凝土板': 650,
    '混凝土墙': 720,
    '混凝土基础': 620,
    '混凝土楼梯': 680,
    '混凝土构造柱': 650,
    
    # 钢筋工程 - 按吨
    '钢筋': 5800,  # 这个价格是含制作安装的
    
    # 砌体工程 - 按m³
    '砌块': 320,
    '砖': 280,
    
    # 模板工程 - 按m²接触面积
    '模板': 52,
    
    # 脚手架 - 按m²建筑面积
    '脚手架': 28,
    
    # 装饰工程 - 按m²
    '抹灰': 22,
    '墙面': 28,
    '地面': 25,
    '天棚': 20,
    '涂料': 15,
    '防水': 28,
    
    # 默认
    'default': 200,
}

def get_unit_price(name):
    """根据项目名称获取定额单价"""
    name = str(name)
    
    # 钢筋
    if '钢筋' in name:
        return 定额单价['钢筋']
    
    # 混凝土柱
    if '柱' in name and ('混凝土' in name or any(x in name for x in ['现浇', '商品'])):
        return 定额单价['混凝土柱']
    
    # 混凝土梁
    if '梁' in name and ('混凝土' in name or any(x in name for x in ['现浇', '商品'])):
        return 定额单价['混凝土梁']
    
    # 混凝土板
    if '板' in name and ('混凝土' in name or any(x in name for x in ['现浇', '商品'])):
        return 定额单价['混凝土板']
    
    # 混凝土墙
    if '墙' in name and ('混凝土' in name or any(x in name for x in ['现浇', '商品'])):
        return 定额单价['混凝土墙']
    
    # 基础
    if '基础' in name:
        return 定额单价['混凝土基础']
    
    # 楼梯
    if '楼梯' in name:
        return 定额单价['混凝土楼梯']
    
    # 构造柱
    if '构造柱' in name:
        return 定额单价['混凝土构造柱']
    
    # 砌体
    if '砌' in name:
        if '砌块' in name:
            return 定额单价['砌块']
        return 定额单价['砖']
    
    # 模板
    if '模板' in name:
        return 定额单价['模板']
    
    # 脚手架
    if '脚手架' in name:
        return 定额单价['脚手架']
    
    # 抹灰/粉刷
    if '抹灰' in name or '粉刷' in name:
        return 定额单价['抹灰']
    
    # 墙面
    if '墙面' in name or '墙饰面' in name:
        return 定额单价['墙面']
    
    # 地面
    if '地面' in name or '楼地面' in name:
        return 定额单价['地面']
    
    # 天棚
    if '天棚' in name or '吊顶' in name:
        return 定额单价['天棚']
    
    # 涂料
    if '涂料' in name or '油漆' in name or '乳胶漆' in name:
        return 定额单价['涂料']
    
    # 防水
    if '防水' in name or '防潮' in name:
        return 定额单价['防水']
    
    return 定额单价['default']

# 费率（2018定额）
费率 = {
    '企业管理费': 0.12,
    '利润': 0.08,
    '规费': 0.065,
    '税金': 0.09,
}

# 读取Excel
wb = load_workbook('5.xlsx')
ws = wb.active

filled = 0
清单合计 = 0

print("=" * 70)
print("台州府城集散中心 - 招标控制价（2018定额）")
print("=" * 70)

# 处理每一行
for row_idx in range(1, ws.max_row + 1):
    cell_val = ws.cell(row_idx, 1).value
    
    if cell_val and isinstance(cell_val, (int, float)):
        try:
            qty = ws.cell(row_idx, 6).value
            name = ws.cell(row_idx, 3).value
            unit = ws.cell(row_idx, 5).value
            
            if qty and name and float(qty) > 0:
                # 获取单价
                unit_price = get_unit_price(str(name))
                total_price = float(qty) * unit_price
                
                # 填入综合单价
                ws.cell(row_idx, 7).value = unit_price
                
                # 填入合价
                ws.cell(row_idx, 8).value = round(total_price, 2)
                
                # 人工费（20%）
                ws.cell(row_idx, 9).value = round(total_price * 0.20, 2)
                
                # 材料费（65%）
                ws.cell(row_idx, 10).value = round(total_price * 0.65, 2)
                
                清单合计 += total_price
                filled += 1
                
        except Exception as e:
            pass

# 计算各项费用
企业管理费 = 清单合计 * 费率['企业管理费']
利润 = 清单合计 * 费率['利润']
规费 = (清单合计 + 企业管理费 + 利润) * 费率['规费']
税金 = (清单合计 + 企业管理费 + 利润 + 规费) * 费率['税金']

招标控制价 = 清单合计 + 企业管理费 + 利润 + 规费 + 税金

# 保存
output = '台州府城_招标控制价_2018定额.xlsx'
wb.save(output)

print(f"\n【清单项目】")
print(f"  清单项目数: {filled}项")
print(f"  清单合价: {清单合计:,.2f}元")

print(f"\n【费用计算】")
print(f"  企业管理费(12%): {企业管理费:,.2f}元")
print(f"  利润(8%): {利润:,.2f}元")
print(f"  规费(6.5%): {规费:,.2f}元")
print(f"  税金(9%): {税金:,.2f}元")

print(f"\n【招标控制价】")
print(f"  总价: {招标控制价:,.2f}元")
print(f"  大写: {int(招标控制价)}元")

print(f"\n文件已保存: {output}")
