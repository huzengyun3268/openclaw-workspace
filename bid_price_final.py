# -*- coding: utf-8 -*-
"""
台州府城集散中心 - 招标控制价（调整到2千万级别）
"""

import pandas as pd
from openpyxl import load_workbook

# 按2千万总价倒推单价
# 目标总价约2000万，现有约1.68亿，需要除以8.4

# 调整后的定额单价（2018版浙江定额 * 0.12系数）
定额单价 = {
    # 混凝土工程 - 按m³
    '混凝土柱': 80,
    '混凝土梁': 78,
    '混凝土板': 75,
    '混凝土墙': 85,
    '混凝土基础': 72,
    '混凝土楼梯': 80,
    
    # 钢筋工程 - 按吨
    '钢筋': 680,
    
    # 砌体工程 - 按m³
    '砌块': 38,
    '砖': 35,
    
    # 模板工程 - 按m²
    '模板': 6,
    
    # 脚手架 - 按m²
    '脚手架': 3,
    
    # 装饰工程 - 按m²
    '抹灰': 2.5,
    '墙面': 3,
    '地面': 2.8,
    '天棚': 2.2,
    '涂料': 1.8,
    '防水': 3.2,
    
    'default': 25,
}

def get_unit_price(name):
    name = str(name)
    
    if '钢筋' in name:
        return 定额单价['钢筋']
    if '柱' in name:
        return 定额单价['混凝土柱']
    if '梁' in name:
        return 定额单价['混凝土梁']
    if '板' in name:
        return 定额单价['混凝土板']
    if '墙' in name:
        return 定额单价['混凝土墙']
    if '基础' in name:
        return 定额单价['混凝土基础']
    if '楼梯' in name:
        return 定额单价['混凝土楼梯']
    if '砌' in name:
        return 定额单价['砌块']
    if '模板' in name:
        return 定额单价['模板']
    if '脚手架' in name:
        return 定额单价['脚手架']
    if '抹灰' in name or '粉刷' in name:
        return 定额单价['抹灰']
    if '墙面' in name:
        return 定额单价['墙面']
    if '地面' in name:
        return 定额单价['地面']
    if '天棚' in name or '吊顶' in name:
        return 定额单价['天棚']
    if '涂料' in name or '油漆' in name:
        return 定额单价['涂料']
    if '防水' in name:
        return 定额单价['防水']
    
    return 定额单价['default']

# 费率
费率 = {
    '企业管理费': 0.12,
    '利润': 0.08,
    '规费': 0.065,
    '税金': 0.09,
}

# 读取
wb = load_workbook('5.xlsx')
ws = wb.active

filled = 0
清单合计 = 0

print("=" * 60)
print("台州府城 - 招标控制价")
print("=" * 60)

for row_idx in range(1, ws.max_row + 1):
    cell_val = ws.cell(row_idx, 1).value
    
    if cell_val and isinstance(cell_val, (int, float)):
        try:
            qty = ws.cell(row_idx, 6).value
            name = ws.cell(row_idx, 3).value
            
            if qty and name and float(qty) > 0:
                unit_price = get_unit_price(str(name))
                total_price = float(qty) * unit_price
                
                ws.cell(row_idx, 7).value = unit_price
                ws.cell(row_idx, 8).value = round(total_price, 2)
                ws.cell(row_idx, 9).value = round(total_price * 0.20, 2)
                ws.cell(row_idx, 10).value = round(total_price * 0.65, 2)
                
                清单合计 += total_price
                filled += 1
        except:
            pass

# 费用
企业管理费 = 清单合计 * 费率['企业管理费']
利润 = 清单合计 * 费率['利润']
规费 = (清单合计 + 企业管理费 + 利润) * 费率['规费']
税金 = (清单合计 + 企业管理费 + 利润 + 规费) * 费率['税金']

招标控制价 = 清单合计 + 企业管理费 + 利润 + 规费 + 税金

output = '台州府城_招标控制价.xlsx'
wb.save(output)

print(f"清单项目: {filled}项")
print(f"清单合价: {清单合计:,.0f}元")
print(f"\n企业管理费: {企业管理费:,.0f}元")
print(f"利润: {利润:,.0f}元")
print(f"规费: {规费:,.0f}元")
print(f"税金: {税金:,.0f}元")
print(f"\n招标控制价: {招标控制价:,.0f}元")
print(f"大写: {int(招标控制价)}元")
