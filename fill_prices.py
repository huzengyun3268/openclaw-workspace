# -*- coding: utf-8 -*-
import pandas as pd
from openpyxl import load_workbook

# 估算单价（2024浙江定额参考）
DINGE_PRICES = {
    '基础': 680, '柱': 720, '梁': 695, '板': 665, '墙': 750,
    '楼梯': 680, '构造柱': 650, '地坪': 350,
    '钢筋': 5800, '箍筋': 5800,
    '砌筑': 380, '砌块': 320, '砖': 350,
    '模板': 58, '脚手架': 35,
    '抹灰': 28, '粉刷': 28, '墙面': 45, '墙饰面': 45,
    '地面': 35, '楼地面': 35, '天棚': 32, '吊顶': 32,
    '涂料': 18, '油漆': 25, '乳胶漆': 22, '防水': 32, '防潮': 32,
    '保温': 65,
    'default': 350,
}

def get_price(name):
    if not isinstance(name, str):
        return DINGE_PRICES['default']
    
    if name in DINGE_PRICES:
        return DINGE_PRICES[name]
    
    if '钢筋' in name or '箍筋' in name:
        return 5800
    if '基础' in name:
        return 680
    if '柱' in name:
        return 720
    if '梁' in name:
        return 695
    if '板' in name:
        return 665
    if '墙' in name:
        return 750
    if '楼梯' in name:
        return 680
    if '构造柱' in name or '构造住' in name:
        return 650
    if '砌' in name:
        return 380
    if '模板' in name or '模扳' in name:
        return 58
    if '脚手架' in name or '脚手扳' in name:
        return 35
    if '抹灰' in name or '粉刷' in name:
        return 28
    if '墙面' in name or '墙饰面' in name:
        return 45
    if '地面' in name or '楼地面' in name:
        return 35
    if '天棚' in name or '吊顶' in name:
        return 32
    if '涂料' in name or '油漆' in name or '乳胶漆' in name:
        return 22
    if '防水' in name or '防潮' in name:
        return 32
    if '保温' in name:
        return 65
    if '地坪' in name or '垫层' in name:
        return 350
    
    return 350

# 读取原Excel
wb = load_workbook('5.xlsx')
ws = wb.active

filled_count = 0
total_amount = 0

print("Processing...")

for row_idx in range(1, ws.max_row + 1):
    cell_val = ws.cell(row_idx, 1).value
    
    if cell_val and isinstance(cell_val, (int, float)):
        try:
            qty = ws.cell(row_idx, 6).value
            name = ws.cell(row_idx, 3).value
            
            if qty and name:
                unit_price = get_price(str(name))
                total = float(qty) * unit_price
                
                # 填入单价（第7列）
                ws.cell(row_idx, 7).value = unit_price
                
                # 填入合价（第8列）
                ws.cell(row_idx, 8).value = round(total, 2)
                
                # 人工费（第9列）- 按30%
                labor = round(total * 0.30, 2)
                ws.cell(row_idx, 9).value = labor
                
                # 材料费（第10列）- 按60%
                material = round(total * 0.60, 2)
                ws.cell(row_idx, 10).value = material
                
                total_amount += total
                filled_count += 1
                
        except:
            pass

# 保存文件
output_file = '台州府城_清单_已填单价.xlsx'
wb.save(output_file)

print(f"\nDone! Filled {filled_count} items")
print(f"Total: {total_amount:,.2f} Yuan")
print(f"File: {output_file}")
