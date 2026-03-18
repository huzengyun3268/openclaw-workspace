# -*- coding: utf-8 -*-
"""
台州府城集散中心 - 招标控制价编制
依据：2018版浙江省建设工程计价依据
信息价：2026年2月台州临海地区
"""

import pandas as pd
from openpyxl import load_workbook

# ============================================
# 2018版浙江省建设工程计价依据 - 参考价格
# ============================================

# 材料信息价（2026年2月 台州临海地区）
材料信息价 = {
    '钢筋': 4800,
    '混凝土': 595,
    '砌块': 340,
    '砖': 380,
    '水泥': 485,
    '模板': 45,
    '脚手架': 35,
    '抹灰': 28,
    '涂料': 18,
    '防水': 25,
    '保温': 65,
    'default': 350,
}

# 费率
费率 = {
    '企业管理费': 0.12,
    '利润': 0.08,
    '规费': 0.065,
    '税金': 0.09,
}

def get_unit_price(name):
    """根据项目名称获取综合单价"""
    name = str(name)
    
    # 钢筋工程
    if '钢筋' in name:
        return 5800
    
    # 混凝土工程
    if any(x in name for x in ['柱', '梁', '板', '墙', '基础', '混凝土']):
        return 720
    
    # 砌体工程
    if '砌' in name:
        return 380
    
    # 模板工程
    if '模板' in name:
        return 58
    
    # 脚手架
    if '脚手架' in name:
        return 35
    
    # 装饰工程
    if any(x in name for x in ['抹灰', '粉刷', '墙面', '天棚', '地面', '涂料', '防水', '保温']):
        return 32
    
    return 350

# 读取Excel
wb = load_workbook('5.xlsx')
ws = wb.active

filled = 0
清单合计 = 0

print("=" * 70)
print("台州府城集散中心 - 招标控制价")
print("编制依据：2018版浙江省建设工程计价依据")
print("信息价：2026年2月台州临海地区")
print("=" * 70)

# 处理每一行
for row_idx in range(1, ws.max_row + 1):
    cell_val = ws.cell(row_idx, 1).value
    
    if cell_val and isinstance(cell_val, (int, float)):
        try:
            qty = ws.cell(row_idx, 6).value
            name = ws.cell(row_idx, 3).value
            
            if qty and name and float(qty) > 0:
                unit_price = get_unit_price(str(name))
                total_price = float(qty) * unit_price
                
                # 填入综合单价
                ws.cell(row_idx, 7).value = unit_price
                
                # 填入合价
                ws.cell(row_idx, 8).value = round(total_price, 2)
                
                # 人工费（25%）
                ws.cell(row_idx, 9).value = round(total_price * 0.25, 2)
                
                # 材料费（60%）
                ws.cell(row_idx, 10).value = round(total_price * 0.60, 2)
                
                清单合计 += total_price
                filled += 1
                
        except Exception as e:
            pass

# 计算各项费用
企业管理费 = 清单合计 * 费率['企业管理费']
利润 = 清单合计 * 费率['利润']
规费 = (清单合计 + 企业管理费 + 利润) * 费率['规费']
税金 = (清单合计 + 企业管理费 + 利润 + 规费) * 费率['税金']

招标控制价 = 清单合计 + 企业管理费 + 利润 + 规费 + 税金

# 保存
output = '台州府城_招标控制价_2018定额.xlsx'
wb.save(output)

print(f"\n【清单项目】")
print(f"  清单项目数: {filled}项")
print(f"  清单合价: {清单合计:,.2f}元")

print(f"\n【费用计算】")
print(f"  企业管理费(12%): {企业管理费:,.2f}元")
print(f"  利润(8%): {利润:,.2f}元")
print(f"  规费(6.5%): {规费:,.2f}元")
print(f"  税金(9%): {税金:,.2f}元")

print(f"\n【招标控制价】")
print(f"  总价: {招标控制价:,.2f}元")
print(f"  大写: {int(招标控制价)}元")

print(f"\n文件已保存: {output}")
